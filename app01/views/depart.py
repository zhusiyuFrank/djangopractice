from openpyxl import load_workbook

from django import forms
from django.shortcuts import render, redirect
from django.core.exceptions import ValidationError

from app01 import models
from app01.utils.pagination import Pagination


# Create your views here.
def depart_list(request):
    data_dict = {}
    value = request.GET.get('q', "")
    if value:
        data_dict["title__contains"] = value  # 含有用户输入的value的title
    departList = models.Department.objects.filter(**data_dict)

    page_obj = Pagination(request, departList)

    context = {
        'departList': page_obj.page_queryset,
        'value': value,
        'page_string': page_obj.html()
    }
    return render(request, 'depart_list.html', context)


class DepartModelForm(forms.ModelForm):
    class Meta:
        model = models.Department
        fields = ["title"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}

    def clean_title(self):
        depart_name = self.cleaned_data["title"]
        exists = models.Department.objects.filter(title=depart_name).exists()
        if exists:
            raise ValidationError("部门已存在")
        return depart_name


def depart_add(request):
    if request.method == 'GET':
        form = DepartModelForm()
        return render(request, 'depart_add.html', {'form': form})
    else:
        form = DepartModelForm(data=request.POST)
        if form.is_valid():
            form.save()
            return redirect('/depart/list/')
        else:
            return render(request, 'depart_add.html', {'form': form})


def depart_delete(request):
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


class DepartEditModelForm(forms.ModelForm):
    class Meta:
        model = models.Department
        fields = ["title"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control"}

    def clean_title(self):
        depart_name = self.cleaned_data["title"]
        exists = models.Department.objects.exclude(id=self.instance.pk).filter(title=depart_name).exists()
        if exists:
            raise ValidationError("部门已存在")
        return depart_name


def depart_edit(request, nid):
    row_obj = models.Department.objects.filter(id=nid).first()
    if request.method == 'GET':
        form = DepartEditModelForm(instance=row_obj)
        return render(request, 'depart_edit.html', {'form': form})
    else:
        form = DepartEditModelForm(data=request.POST, instance=row_obj)
        if form.is_valid():
            form.save()
            return redirect('/depart/list/')
        else:
            return render(request, 'depart_edit.html', {'form': form})


def depart_upload(request):
    if request.method == 'GET':
        return render(request, 'depart_upload.html')
    else:
        # 获取用户上传文件对象
        file_object = request.FILES.get("exc")
        # 使用openpyxl读取文件内容
        wb = load_workbook(file_object)
        # excel文件中第一个sheet
        sheet = wb.worksheets[0]
        # 从第二行开始读取
        for row in sheet.iter_rows(min_row=2):
            text = row[0].value
            exists = models.Department.objects.filter(title=text).exists()
            if not exists:
                models.Department.objects.create(title=text)
        return redirect('/depart/list/')
